#! python3
# creates new xlsx file with multiplication table on first sheet in desired size
# multiplicationTable (dim)

import openpyxl
from openpyxl.styles import Font
dim = 8 #define size of a table

wb = openpyxl.Workbook()
ws = wb.active
c = 0
r = 0
for r in range(dim + 1):
    r +=1
    c +=1
    ws.cell(r + 1,1).value = r
    ws.cell(r + 1,1).font = Font(bold=True)
    ws.cell(1,c + 1).value = c
    ws.cell(1,c+1).font = Font(bold=True)
for r in range(dim + 1):
    r +=1 
    for c in range(dim + 1):
        c+=1
        print(c)
        print(r * c)
        ws.cell(r + 1,c + 1).value = r * c

wb.save('multiplicationTable.xlsx')



