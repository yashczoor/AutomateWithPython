#! python3
#transposes a table in workbook active sheet and saves to new worksheet

import openpyxl

wb = openpyxl.load_workbook('transposeTestFile.xlsx')
ws = wb.active
sheetData = []
for c in range(1, ws.max_column + 1):
    col = []
    for r in range(1, ws.max_row + 1):
        x = r - 1
        y = c - 1
        col.append(ws.cell(r,c).value)
    sheetData.append(col)

for row in sheetData:
    print(row)

wsTarget = wb.create_sheet('transposed')
for c in range(1, ws.max_column + 1):
    col = []
    for r in range(1, ws.max_row + 1):
        x = r - 1
        y = c - 1
        wsTarget.cell(c,r).value = sheetData[y][x]

wb.save('transposedData.xlsx')
